import mimetypes

from django.http import HttpResponse
from tablib import Dataset
from django.contrib import messages
from django.db import transaction
from django.shortcuts import render

# Create your views here.
from mysports.views import getUser
from product.forms import CategoryForm, ProductForm, TaxForm, BrandForm, TaxSlabForm
from product.models import ProductClass, Product, ProductCategory, Brand, Tax, TaxSlab
from product.resources import ProductClassResource
from utility.constants import INFO_USERNAME_PWD_RECOVERY, INFO_CLASS_ADDITION_SUCCESS, INFO_CLASS_ADDITION_FAILURE, \
    INFO_CLASS_INACTIVATION_FAIL, INFO_CLASS_INACTIVATION_SUCCESS, INFO_CATEGORY_ADDITION_SUCCESS, \
    INFO_CATEGORY_ADDITION_FAILURE, INFO_PRODUCT_ADDITION_FAILURE, INFO_PRODUCT_ADDITION_SUCCESS, \
    INFO_PRODUCT_INACTIVATION_SUCCESS, INFO_PRODUCT_INACTIVATION_FAIL, INFO_CATEGORY_INACTIVATION_FAIL, \
    INFO_CATEGORY_INACTIVATION_SUCCESS
import openpyxl
import os
from django.conf import settings

def addProductClass(request):
    context = {}
    if request.method == 'POST':
        login_data = request.POST.dict()
        if request.session.get('userId') is not None:
            userId = int(request.session['userId'])
            try:
                user = getUser(userId)
                with transaction.atomic():
                    productClass = None
                    isActive = ''
                    if login_data['active']:
                        isActive = True
                    else:
                        isActive= False
                    if user != None:
                        name = login_data['className']
                        description = login_data['classDescription']
                        productClass = ProductClass(
                        name=name,
                        classDescription=description,
                        isActive=isActive,
                        createdBy=user
                        )
                        productClass.save()
                        messages.success(request, INFO_CLASS_ADDITION_SUCCESS)
                        listOfClasses = ProductClass.objects.filter(createdBy=user).filter(isActive=True)
                        context['listOfClasses'] = listOfClasses
                        if 'save_display' in request.POST:
                            return render(request, 'product/listProductClass.html', context)
                        elif 'save_add' in request.POST:
                            return render(request, 'product/addProductClass.html', context)
            except Exception as ex:
                print(ex)
                messages.success(request, INFO_CLASS_ADDITION_FAILURE)
                return render(request, 'product/addProductClass.html', context)
    return render(request, 'product/addProductClass.html', context)

def addProductCategory(request):
    context = {}
    if request.method == 'POST':
        login_data = request.POST.dict()
        print(login_data)
        if request.session.get('userId') is not None:
            userId = int(request.session['userId'])
            try:
                user = getUser(userId)
                productClass = ProductClass.objects.get(id=int(login_data['productClass']))
                # productClass = ProductClass.objects.get(id=3)
                print('1')
                with transaction.atomic():
                    productCategory = None
                    isActive = ''
                    if login_data['active']:
                        isActive = True
                    else:
                        isActive = False
                    print('3')
                    if user != None:
                        name = login_data['name']
                        description = login_data['description']
                        productCategory = ProductCategory(
                            name=name,
                            categoryDescription=description,
                            productClass = productClass,
                            isActive=isActive,
                            createdBy=user
                        )
                        productCategory.save()
                        messages.success(request, INFO_CATEGORY_ADDITION_SUCCESS)
                        listOfCategory = ProductCategory.objects.filter(createdBy=user).filter(isActive=True)
                        context['listOfCategory'] = listOfCategory
                        if 'save_display' in request.POST:
                            print('Here')
                            return render(request, 'product/listProductCategory.html', context)
                        elif 'save_add' in request.POST:
                            print('Here 1')
                            form1 = CategoryForm(request.POST or None)
                            context = {
                                "form1": form1
                            }
                            return render(request, 'product/addProductCategory.html', context)
            except Exception as ex:
                print(ex)
                messages.success(request, INFO_CATEGORY_ADDITION_FAILURE)
                form1 = CategoryForm(request.POST or None)
                context = {
                    "form1": form1
                }

                return render(request, 'product/addProductCategory.html', context)
    form1 = CategoryForm(request.POST or None)
    # form1 = CategoryForm()
    # if form1.is_valid():
    #     instance = form1.save(commit=False)
    #     instance.save()

    context = {
        "form1": form1
    }
    return render(request, 'product/addProductCategory.html', context)

def addProduct(request):
    context = {}
    if request.method == 'POST':
        login_data = request.POST.dict()
        print(login_data)
        if request.session.get('userId') is not None:
            userId = int(request.session['userId'])
            try:
                user = getUser(userId)
                productCategory = ProductCategory.objects.get(id=int(login_data['Category']))
                # productClass = ProductClass.objects.get(id=3)
                # imageFile = request.FILES['imageFile']
                with transaction.atomic():
                    product = None
                    isActive = True
                    if 'active' in request.POST:
                        isActive = True
                    else:
                        isActive = False
                    discountAvailable = False
                    if 'isDiscountAvailable' in request.POST:
                        discountAvailable = True
                    else:
                        discountAvailable = False
                    manufactureDate = '9999-12-31'
                    if login_data['manufactureDate']:
                        manufactureDate = login_data['manufactureDate'],
                        manufactureDate = manufactureDate[0]

                    expiryDate = '9999-12-31'
                    if login_data['expiryDate']:
                        expiryDate = login_data['expiryDate'],
                        expiryDate = expiryDate[0]

                    if user != None:
                        name = login_data['name']
                        description = login_data['description']
                        productSpecification = login_data['productSpecification']
                        brand = Brand.objects.get(id=int(login_data['brand']))
                        MRP = login_data['mrp']
                        discountPercentage = login_data['discount']
                        if discountPercentage == '' or discountPercentage == None:
                            discountPercentage = 0
                        batchNumber = login_data['batchNumber']
                        tax = Tax.objects.get(id=int(login_data['tax']))
                        taxSlab = TaxSlab.objects.get(id=int(login_data['TaxSlab']))
                        # manufactureDate = login_data['manufactureDate']
                        # expiryDate = login_data['expiryDate']
                        sellingPrice = int(MRP) - (int(MRP)*int(discountPercentage)/100)
                        print('I am active' , isActive)
                        product = Product(
                            name=name,
                            prodDescription=description,
                            prodCategoryId = productCategory,
                            isActive=isActive,
                            brand = brand,
                            MRP = MRP,
                            sellingPrice = sellingPrice,
                            isDiscountAvailable = discountAvailable,
                            discountPercentage = discountPercentage,
                            productSpecification = productSpecification,
                            batchNumber = batchNumber,
                            manufactureDate = manufactureDate,
                            tax = tax,
                            taxSlab = taxSlab,
                            expiryDate = expiryDate,
                            createdBy = user,
                        )
                        product.save()
                        messages.success(request, INFO_PRODUCT_ADDITION_SUCCESS)
                        listOfProducts = Product.objects.filter(createdBy=user).filter(isActive=True)
                        context['listOfProducts'] = listOfProducts
                        if 'save_display' in request.POST:
                            print('Here')
                            return render(request, 'product/listOfProducts.html', context)
                        elif 'save_add' in request.POST:
                            print('Here 1')
                            addProductsInLineForm(request, context)
                            return render(request, 'product/addProduct.html', context)
            except Exception as ex:
                print(ex)
                messages.error(request, INFO_PRODUCT_ADDITION_FAILURE)
                addProductsInLineForm(request, context)
                return render(request, 'product/addProduct.html', context)
    # form1 = ProductForm(request.POST or None)
    # # form1 = CategoryForm()
    # if form1.is_valid():
    #     instance = form1.save(commit=False)
    #     instance.save()
    #
    # form2 = TaxForm(request.POST or None)
    # # form1 = CategoryForm()
    # if form2.is_valid():
    #     instance = form2.save(commit=False)
    #     instance.save()
    #
    # form3 = BrandForm(request.POST or None)
    # # form1 = CategoryForm()
    # if form3.is_valid():
    #     instance = form3.save(commit=False)
    #     instance.save()
    #
    # form4 = TaxSlabForm(request.POST or None)
    # # form1 = CategoryForm()
    # if form4.is_valid():
    #     instance = form4.save(commit=False)
    #     instance.save()
    #
    # context['form1'] = form1
    # context['form2'] = form2
    # context['form3'] = form3
    # context['form4'] = form4
    addProductsInLineForm(request,context)
    return render(request, 'product/addProduct.html', context)

def addProductsInLineForm(request,context):
    form1 = ProductForm(request.POST or None)
    # form1 = CategoryForm()
    # if form1.is_valid():
    #     instance = form1.save(commit=False)
    #     instance.save()

    form2 = TaxForm(request.POST or None)
    # form1 = CategoryForm()
    # if form2.is_valid():
    #     instance = form2.save(commit=False)
    #     instance.save()

    form3 = BrandForm(request.POST or None)
    # form1 = CategoryForm()
    # if form3.is_valid():
    #     instance = form3.save(commit=False)
    #     instance.save()

    form4 = TaxSlabForm(request.POST or None)
    # form1 = CategoryForm()
    # if form4.is_valid():
    #     instance = form4.save(commit=False)
    #     instance.save()

    context['form1'] = form1
    context['form2'] = form2
    context['form3'] = form3
    context['form4'] = form4
def listProductClass(request, userId):
    context = {}
    user = getUser(userId)
    listOfClasses = ProductClass.objects.filter(createdBy=user).filter(isActive=True)
    print(listOfClasses)
    context['listOfClasses'] = listOfClasses
    return render(request, 'product/listProductClass.html', context)

def listProductCategory(request, userId):
    context = {}
    user = getUser(userId)
    listOfCategory = ProductCategory.objects.filter(createdBy=user).filter(isActive=True)
    print(listOfCategory)
    context['listOfCategory'] = listOfCategory
    print(listOfCategory)
    return render(request, 'product/listProductCategory.html', context)

def listProductCategoryClassWise(request, userId):
    context = {}
    user = getUser(userId)
    listOfCategory = ProductCategory.objects.filter(createdBy=user).filter(isActive=True).order_by('productClass')
    print(listOfCategory)
    context['listOfCategory'] = listOfCategory
    print(listOfCategory)
    return render(request, 'product/listProductCategoryClassWise.html', context)


def listProduct(request, userId):
    context = {}
    user = getUser(userId)
    listOfProducts = Product.objects.filter(createdBy=user).filter(isActive=True)
    print(listOfProducts)
    context['listOfProducts'] = listOfProducts
    print(listOfProducts)
    return render(request, 'product/listOfProducts.html', context)


def inactivateProductClass(request, classId,userId):
    context = {}
    user = getUser(userId)
    try:
        if classId == 0:
            productClasses = ProductClass.objects.all().filter(createdBy=user)
            productClasses.update(isActive=False)
        else:
            productClass = ProductClass.objects.get(id=classId)
            productClass.isActive = False
            productClass.save()
        listOfClasses = ProductClass.objects.filter(createdBy=user).filter(isActive=True)
        print(listOfClasses)
        context['listOfClasses'] = listOfClasses
        messages.success(request, INFO_CLASS_INACTIVATION_SUCCESS)
        return render(request, 'product/listProductClass.html', context)
    except Exception as ex:
        print(ex)
        listOfClasses = ProductClass.objects.filter(createdBy=user).filter(isActive=True)
        context['listOfClasses'] = listOfClasses
        messages.error(request, INFO_CLASS_INACTIVATION_FAIL)
        return render(request, 'product/listProductClass.html', context)

def inactivateProductCategory(request, categoryId,userId):
    context = {}
    user = getUser(userId)
    try:
        if categoryId == 0:
            productCategory = ProductCategory.objects.all().filter(createdBy=user)
            productCategory.update(isActive=False)
        else:
            productCategory = ProductCategory.objects.get(id=categoryId)
            productCategory.isActive = False
            productCategory.save()
        listOfCategory = ProductCategory.objects.filter(createdBy=user).filter(isActive=True)
        print(listOfCategory)
        context['listOfCategory'] = listOfCategory
        messages.success(request, INFO_CATEGORY_INACTIVATION_SUCCESS)
        return render(request, 'product/listProductCategory.html', context)
    except Exception as ex:
        print(ex)
        listOfCategory = ProductCategory.objects.filter(createdBy=user).filter(isActive=True)
        context['listOfCategory'] = listOfCategory
        messages.error(request, INFO_CATEGORY_INACTIVATION_FAIL)
        return render(request, 'product/listProductCategory.html', context)

def inactivateProduct(request, productId,userId):
    context = {}
    user = getUser(userId)
    try:
        if productId == 0:
            products = Product.objects.all().filter(createdBy=user)
            products.update(isActive=False)
        else:
            product = Product.objects.get(id=productId)
            product.isActive = False
            product.save()
        listOfProducts = Product.objects.filter(createdBy=user).filter(isActive=True)
        print(listOfProducts)
        context['listOfProducts'] = listOfProducts
        messages.success(request, INFO_PRODUCT_INACTIVATION_SUCCESS)
        return render(request, 'product/listOfProducts.html', context)
    except Exception as ex:
        print(ex)
        listOfProducts = Product.objects.filter(createdBy=user).filter(isActive=True)
        context['listOfProducts'] = listOfProducts
        messages.error(request, INFO_PRODUCT_INACTIVATION_FAIL)
        return render(request, 'product/listOfProducts.html', context)

def bulkUploadProductClass(request):
    context= {}
    if request.method == 'POST':
        excel_file = request.FILES['myfile']
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["PCLASS"]
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            print(row)
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        print(len(excel_data))
        # Since first element is going to be only the header part
        excel_data.pop(0)
        print(len(excel_data))
        if request.session.get('userId') is not None:
            userId = int(request.session['userId'])
            try:
                user = getUser(userId)
                # with transaction.atomic():
                productClass = None
                isActive = ''
                if user != None:
                    for row in excel_data:
                        name = row[0]
                        description = row[1]
                        isActive = row[2]
                        print(name, description , isActive)
                        productClass = ProductClass(
                        name=name,
                        classDescription=description,
                        isActive=isActive,
                        createdBy=user
                        )
                        productClass.save()
                    messages.success(request, INFO_CLASS_ADDITION_SUCCESS)
                    listOfClasses = ProductClass.objects.filter(createdBy=user).filter(isActive=True)
                    context['listOfClasses'] = listOfClasses
                    return render(request, 'product/listProductClass.html', context)
            except Exception as ex:
                print(ex)
                messages.success(request, INFO_CLASS_ADDITION_FAILURE)
                return render(request, 'product/bulkUploadProductClass.html')
    else:
        return render(request, 'product/bulkUploadProductClass.html')

def bulkUploadProductCategory(request):
    context= {}
    if request.method == 'POST':
        excel_file = request.FILES['myfile']
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["PCAT"]
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            print(row)
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        print(len(excel_data))
        # Since first element is going to be only the header part
        excel_data.pop(0)
        print(len(excel_data))
        if request.session.get('userId') is not None:
            userId = int(request.session['userId'])
            try:
                user = getUser(userId)
                # with transaction.atomic():
                productClass = None
                isActive = ''
                if user != None:
                    for row in excel_data:
                        name = row[0]
                        description = row[1]
                        isActive = row[2]
                        productClass = row[3]
                        print(name, description , isActive,productClass)
                        productClass = ProductClass.objects.get(id=int(productClass))
                        productCategory = ProductCategory(
                        name=name,
                        categoryDescription=description,
                        isActive=isActive,
                        productClass = productClass,
                        createdBy=user
                        )
                        productCategory.save()
                    messages.success(request, INFO_CATEGORY_ADDITION_SUCCESS)
                    listOfCategory = ProductCategory.objects.filter(createdBy=user).filter(isActive=True)
                    context['listOfCategory'] = listOfCategory
                    return render(request, 'product/listProductCategory.html', context)
            except Exception as ex:
                print(ex)
                messages.success(request, INFO_CATEGORY_ADDITION_FAILURE)
                return render(request, 'product/bulkUploadProductClass.html')
    else:
    #
    #     # file_path = os.path.join(settings.BASE_DIR, 'relative_path')
    #     relative_path = '/product/productCategory.xlsx'
    #     file_path = os.path.join(settings.BASE_DIR,relative_path )
    #     print(file_path)
    #     # print(settings.BASE_DIR,'product/templates/product/productCategory.xlsx')
    #     context['file_path'] = relative_path
        return render(request, 'product/bulkUploadProductCategory.html',context)

def bulkUploadProducts(request):
    context= {}
    if request.method == 'POST':
        excel_file = request.FILES['myfile']
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        worksheet = wb["PRODUCT"]
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            print(row)
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        print(len(excel_data))
        # Since first element is going to be only the header part
        excel_data.pop(0)
        print(len(excel_data))
        if request.session.get('userId') is not None:
            userId = int(request.session['userId'])
            try:
                user = getUser(userId)
                # with transaction.atomic():
                isActive = ''
                if user != None:
                    for row in excel_data:
                        name = row[0]
                        print(name)
                        prodDescription = row[1]
                        print(prodDescription)
                        brand = row[2]
                        print(brand)
                        prodCategoryId = row[3]
                        print(prodCategoryId)
                        MRP = row[4]
                        print(MRP)
                        isDiscountAvailable = row[5]
                        print(isDiscountAvailable)
                        discountPercentage = row[6]
                        print(discountPercentage)
                        active = row[7]
                        print(active)
                        productSpecification = row[8]
                        print(productSpecification)
                        batchNumber = row[9]
                        print(batchNumber)
                        manufactureDate = row[10]
                        print(manufactureDate)
                        tax = row[11]
                        print(tax)
                        taxSlab = row[12]
                        print(taxSlab)
                        expiryDate = row[13]
                        print(expiryDate)
                        if active=='TRUE' or active=='True':
                            active = True
                        else:
                            active = False
                        discountAvailable = False
                        if isDiscountAvailable == 'TRUE':
                            discountAvailable = True
                        else:
                            discountAvailable = False
                        sellingPrice = int(MRP) - (int(MRP) * int(discountPercentage) / 100)

                        brand = Brand.objects.get(id=int(brand))
                        prodCategoryId = ProductCategory.objects.get(id=int(prodCategoryId))
                        tax = Tax.objects.get(id=int(tax))
                        taxSlab = TaxSlab.objects.get(id=int(taxSlab))
                        print('Active',active)
                        print('Discount available',discountAvailable)
                        product = Product(
                            name=name,
                            prodDescription=prodDescription,
                            prodCategoryId=prodCategoryId,
                            isActive=active,
                            brand=brand,
                            MRP=MRP,
                            sellingPrice=sellingPrice,
                            isDiscountAvailable=discountAvailable,
                            discountPercentage=discountPercentage,
                            productSpecification=productSpecification,
                            batchNumber=batchNumber,
                            manufactureDate=manufactureDate,
                            tax=tax,
                            taxSlab=taxSlab,
                            expiryDate=expiryDate,
                            createdBy=user,
                        )
                        product.save()
                    messages.success(request, INFO_PRODUCT_ADDITION_SUCCESS)
                    listOfProducts = Product.objects.filter(createdBy=user).filter(isActive=True)
                    context['listOfProducts'] = listOfProducts
                    return render(request, 'product/listOfProducts.html', context)
            except Exception as ex:
                print(ex)
                messages.success(request, INFO_PRODUCT_ADDITION_FAILURE)
                return render(request, 'product/addProductsInBulkClass.html')
    else:
        return render(request, 'product/addProductsInBulkClass.html',context)


def download_file(request,filePath):
    # fill these variables with real values
    # fl_path = ‘/file/path'
    filename = 'downloaded_file_name.extension'

    fl = open(filePath, 'rw')
    mime_type, _ = mimetypes.guess_type(filePath)
    response = HttpResponse(fl, content_type=mime_type)
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    return response




def searchProductClass(request,userId):
    context = {}
    user = getUser(userId)
    login_data = request.POST.dict()
    print(login_data)
    search_item = login_data['search_item']
    listOfClasses = ProductClass.objects.filter(createdBy=user).filter(isActive=True).filter(name__icontains=search_item)
    print(listOfClasses)
    context['listOfClasses'] = listOfClasses
    return render(request, 'product/listProductClass.html', context)

def searchProductCategory(request,userId):
    context = {}
    user = getUser(userId)
    login_data = request.POST.dict()
    print(login_data)
    search_item = login_data['search_item']
    listOfCategory = ProductCategory.objects.filter(createdBy=user).filter(isActive=True).filter(name__icontains=search_item)
    print(listOfCategory)
    context['listOfCategory'] = listOfCategory
    return render(request, 'product/listProductCategory.html', context)

def searchProduct(request,userId):
    context = {}
    user = getUser(userId)
    login_data = request.POST.dict()
    print(login_data)
    search_item = login_data['search_item']
    listOfProducts = Product.objects.filter(createdBy=user).filter(isActive=True).filter(name__icontains=search_item)
    print(listOfProducts)
    context['listOfProducts'] = listOfProducts
    return render(request, 'product/listOfProducts.html', context)
